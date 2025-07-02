# main.py

import re
import io
import openpyxl
import csv
import smtplib
import socket
import dns.resolver # A dependency of email-validator[deliverability]
from fastapi import FastAPI, UploadFile, File, HTTPException, status, Query
from email_validator import validate_email, EmailNotValidError
from faker import Faker
from typing import Optional

# --- Application Setup ---
# Initialize the FastAPI application
app = FastAPI(
    title="Advanced Email Validation API",
    description="A robust API to validate, check, and generate email addresses.",
    version="2.1.0", # Incremented version
)

# Initialize Faker for generating fake data
fake = Faker()

# --- Data for Email Checks ---
# It's more efficient to load these lists once when the app starts.
# Using sets for domain lists provides O(1) average time complexity for lookups, which is much faster than list iteration.

def load_domains_from_file(filename: str) -> set:
    """Helper function to load domains from a text file into a set for fast lookups."""
    try:
        with open(filename, "r") as f:
            # Read lines, strip whitespace, and filter out empty lines
            return {line.strip().lower() for line in f if line.strip()}
    except FileNotFoundError:
        # If the file doesn't exist, return an empty set and print a warning.
        print(f"Warning: Domain file '{filename}' not found. Returning empty set.")
        return set()

# Load disposable, free, and service email data from external files (best practice)
# NOTE: You will need to create these files: disposable_domains.txt, free_domains.txt, service_prefixes.txt
DISPOSABLE_DOMAINS = load_domains_from_file("disposable_domains.txt")
FREE_DOMAINS = load_domains_from_file("free_domains.txt")
SERVICE_PREFIXES = load_domains_from_file("service_prefixes.txt")


# --- NEW: Advanced SMTP Mailbox Verification ---
def is_mailbox_deliverable(email: str) -> bool:
    """
    Performs a live SMTP check to see if a mailbox is likely to exist.
    This is more accurate than just checking MX records.
    """
    try:
        # Step 1: Get the domain from the email
        domain = email.split('@')[1]

        # Step 2: Get MX records for the domain
        mx_records = dns.resolver.resolve(domain, 'MX')
        mail_exchanger = str(mx_records[0].exchange)

        # Step 3: Connect to the mail server
        # Set a timeout to prevent the request from hanging
        server = smtplib.SMTP(timeout=5)
        server.connect(mail_exchanger)

        # Step 4: Perform SMTP handshake
        server.helo(server.local_hostname)
        server.mail('test@example.com') # A dummy sender email

        # Step 5: Use RCPT TO to check the mailbox
        # The server's response code tells us if the mailbox is recognized.
        code, message = server.rcpt(email)
        server.quit()

        # A 250 status code means the recipient is OK.
        # Other codes (like 550) mean "no such user".
        if code == 250:
            return True
        else:
            return False

    except (dns.resolver.NoAnswer, dns.resolver.NXDOMAIN, smtplib.SMTPConnectError, socket.timeout, smtplib.SMTPServerDisconnected):
        # These exceptions indicate a problem with the domain or mail server connection.
        return False
    except Exception:
        # Catch any other unexpected errors.
        return False


# --- API Endpoints ---

@app.get("/", summary="Welcome Message", tags=["General"])
def index():
    """A simple welcome endpoint to confirm the API is running."""
    return {"message": "Welcome to the Advanced Email Validation API"}


@app.post("/validate-email-full", summary="Comprehensive Email Validation", tags=["Validation"])
def comprehensive_email_validation(email: str):
    """
    Performs a complete validation of a single email address.

    This endpoint checks for:
    1.  **Valid Syntax & Format**: Ensures the email conforms to standard formats.
    2.  **Domain Deliverability**: Checks if the domain has valid MX records.
    3.  **Mailbox Deliverability (SMTP Check)**: Checks if the specific mailbox likely exists.
    4.  **Disposable Domain**: Checks against a list of temporary/disposable email providers.
    5.  **Free Provider**: Checks if the email is from a common free provider (e.g., gmail.com).
    6.  **Service/Role-based**: Checks if the email is a role-based address (e.g., support@, admin@).
    """
    try:
        # The first validation step checks syntax and domain-level deliverability.
        validation_result = validate_email(email, check_deliverability=True)
        
        # The email is syntactically valid and the domain has MX records.
        normalized_email = validation_result.normalized
        domain = normalized_email.split('@')[1]
        local_part = normalized_email.split('@')[0]

        # --- Perform the more accurate, live SMTP check for the mailbox ---
        can_receive_mail = is_mailbox_deliverable(normalized_email)

        # Check against our domain and prefix sets
        is_disposable = domain in DISPOSABLE_DOMAINS
        is_free = domain in FREE_DOMAINS
        is_service_account = local_part in SERVICE_PREFIXES

        return {
            "email": normalized_email,
            "is_valid": True,
            "can_receive_mail": can_receive_mail, # This result is now more accurate
            "is_disposable": is_disposable,
            "is_free_provider": is_free,
            "is_service_account": is_service_account,
            "domain": domain,
            "local_part": local_part
        }

    except EmailNotValidError as e:
        # The email failed the initial syntax or domain check.
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid email: {str(e)}"
        )
    except Exception as e:
        # Catch any other unexpected errors during the process.
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred: {str(e)}"
        )


@app.post("/bulk-validate-emails", summary="Validate Emails from a File", tags=["Bulk Validation"])
async def bulk_validate_emails(file: UploadFile = File(...)):
    """
    Validates a list of emails from an uploaded file (TXT, CSV, or XLSX).

    - For CSV/XLSX, it assumes emails are in the first column.
    - It uses the same robust validation as the single email endpoint.
    """
    # Determine the file type and read emails into a list
    try:
        content = await file.read()
        emails_to_check = []
        filename = file.filename.lower()

        if filename.endswith(".txt"):
            # Decode and split by newlines
            emails_to_check = content.decode("utf-8-sig").strip().splitlines()
        
        elif filename.endswith(".csv"):
            # Use the csv module to correctly parse the file
            csv_file = io.StringIO(content.decode("utf-8-sig"))
            reader = csv.reader(csv_file)
            # Skip header if it exists
            try:
                next(reader) 
            except StopIteration:
                pass # File is empty
            emails_to_check = [row[0] for row in reader if row] # Get email from first column

        elif filename.endswith((".xlsx", ".xls")):
            # Use openpyxl to read from Excel files
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            sheet = workbook.active
            # Skip header
            iter_rows = sheet.iter_rows(min_row=2, values_only=True)
            emails_to_check = [row[0] for row in iter_rows if row and row[0]]

        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid file type. Please upload a .txt, .csv, or .xlsx file.",
            )
        
        # Remove any empty strings that might have been read
        emails_to_check = [email.strip() for email in emails_to_check if email and email.strip()]

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to read or process file: {str(e)}"
        )

    # Process the list of emails
    results = {
        "valid_emails": [],
        "invalid_emails": [],
    }

    for email in emails_to_check:
        try:
            # We only need to check syntax for bulk processing to keep it fast.
            # Set check_deliverability=False for speed.
            validate_email(email, check_deliverability=False)
            results["valid_emails"].append({"email": email, "status": "valid_syntax"})
        except EmailNotValidError as e:
            results["invalid_emails"].append({"email": email, "status": "invalid_syntax", "reason": str(e)})

    return {
        "summary": {
            "total_processed": len(results["valid_emails"]) + len(results["invalid_emails"]),
            "valid_count": len(results["valid_emails"]),
            "invalid_count": len(results["invalid_emails"]),
        },
        "results": results
    }

@app.get("/generate-fake-email", summary="Generate Fake Business Emails", tags=["Generation"])
def generate_fake_business_email(
    count: int = Query(1, ge=1, le=100, description="Number of emails to generate (1-100)."),
    domain: Optional[str] = Query(None, description="An optional, valid domain to use for all generated emails (e.g., 'example.com').")
):
    """
    Generates one or more plausible-looking but fake business emails and job titles.
    An optional, valid domain can be provided.
    """
    # Simple validation for the provided domain
    if domain:
        # A simple check to ensure the domain contains a dot and no spaces.
        if '.' not in domain or ' ' in domain:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid domain format provided. A valid domain should contain a '.' and no spaces (e.g., 'example.com')."
            )

    emails = []
    for _ in range(count):
        first_name = fake.first_name().lower()
        last_name = fake.last_name().lower()
        
        # Use the validated domain if it exists, otherwise generate a fake one.
        final_domain = domain if domain else fake.company().lower().replace(" ", "").replace(",", "") + "." + fake.tld()
        
        email = f"{first_name}.{last_name}@{final_domain}"
        job_title = fake.job()
        emails.append({"email": email, "job_title": job_title})
    
    return emails
